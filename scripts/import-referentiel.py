#!/usr/bin/env python3
"""Genere supabase/seed/referentiel_2026_2030.sql a partir du classeur du referentiel CAPEX.
Usage : python3 scripts/import-referentiel.py docs/data/SEDIMA_Referentiel_Investissements_2026-2030_v4.xlsx
Le SQL produit est idempotent : projets et taches sont mis a jour sur leur code."""
import sys, calendar, datetime as dt
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "docs/data/SEDIMA_Referentiel_Investissements_2026-2030_v4.xlsx"
OUT = "supabase/seed/referentiel_2026_2030.sql"

CATEGORY = {"Œufs de table": "oeufs_table", "Oeufs de table": "oeufs_table", "OAC & Poussins": "oac_poussins",
            "Poulets de chair": "poulet_chair", "Poulet de chair": "poulet_chair", "Industriel": "industriels",
            "Industriels": "industriels", "Autres": "autres"}
STATUS = {"Plan": "plan", "Cadrage": "cadrage", "Approuvé": "approuve", "Engagé": "engage",
          "En exécution": "execution", "Clôturé": "cloture", "Hors périmètre": "hors_perimetre"}

def q(v):
    if v is None: return "null"
    return "'" + str(v).replace("'", "''") + "'"
def n(v):
    return "0" if v in (None, "") else str(round(float(v), 2))
def month_start(s):  # '2026-01' -> 2026-01-01
    y, m = map(int, str(s).split("-")[:2]); return dt.date(y, m, 1)
def month_end(s):
    y, m = map(int, str(s).split("-")[:2]); return dt.date(y, m, calendar.monthrange(y, m)[1])

wb = openpyxl.load_workbook(SRC, data_only=True)
def table(ws, header_row=5):
    hdr = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in r): continue
        yield dict(zip(hdr, r))

projects = [p for p in table(wb["Portefeuille"]) if p.get("Code")]
tasks = [t for t in table(wb["Tâches"]) if t.get("Code projet")]
tranches = [t for t in table(wb["Tranches"]) if t.get("Code")]
years = [int(y) for y in next(wb["Tranches"].iter_rows(min_row=5, max_row=5, values_only=True)) if str(y).isdigit()]

out = ["-- Genere par scripts/import-referentiel.py depuis " + SRC.split("/")[-1],
       "-- Executer dans Supabase > SQL Editor apres 20260905_step2b_import_model.sql. Idempotent.",
       "begin;", ""]

# ---- Projets ----
out.append("-- Projets (28) : mise a jour sur le code")
for p in projects:
    start, end = month_start(p["Début"]), month_end(p["Fin"])
    manager = (p.get("Chef de projet") or "").strip()
    manager_name = "" if manager in ("", "À définir") else manager
    out.append(f"""insert into public.projects (code, name, description, status, category, start_date, end_date, budget, currency,
  site, business_unit, pillar, plan_objective, wbs_nature, data_quality, budget_kpmg, manager_name, source_note, comment, manager_id)
values ({q(p["Code"])}, {q(p["Nom du projet"])}, {q(p.get("Objectif du plan") or "")}, {q(STATUS[p["Statut"]])}, {q(CATEGORY[p["Bucket"]])},
  '{start}', '{end}', {n(p["Enveloppe approuvée"])}, 'XOF',
  {q(p.get("Site") or "")}, {q(p.get("BU") or "")}, {q(p.get("Pilier") or "")}, {q(p.get("Objectif du plan") or "")}, {q(p.get("Nature (modèle WBS)") or "")},
  {q(p.get("Qualité") or "")}, {n(p.get("Enveloppe KPMG"))}, {q(manager_name)}, {q(p.get("Source") or "")}, {q(p.get("Commentaire") or "")},
  (select id from public.profiles where full_name ilike {q(manager_name)} limit 1))
on conflict (code) do update set name = excluded.name, description = excluded.description, status = excluded.status, category = excluded.category,
  start_date = excluded.start_date, end_date = excluded.end_date, budget = excluded.budget, site = excluded.site, business_unit = excluded.business_unit,
  pillar = excluded.pillar, plan_objective = excluded.plan_objective, wbs_nature = excluded.wbs_nature, data_quality = excluded.data_quality,
  budget_kpmg = excluded.budget_kpmg, manager_name = excluded.manager_name, source_note = excluded.source_note, comment = excluded.comment,
  manager_id = coalesce(public.projects.manager_id, excluded.manager_id);""")

# ---- Tranches ----
out.append("\n-- Repartition pluriannuelle")
for t in tranches:
    for y in years:
        out.append(f"insert into public.project_budget_years (project_id, year, amount) select id, {y}, {n(t.get(y) if y in t else t.get(str(y)))} from public.projects where code = {q(t['Code'])} on conflict (project_id, year) do update set amount = excluded.amount;")

# ---- Lots (taches parentes) puis taches, en listes VALUES compactes ----
lots = {}
for t in tasks:
    key = (t["Code projet"], t["Code lot"])
    if key not in lots:
        lots[key] = {"name": t["Nom du lot"], "start": t["Début (sem.)"], "end": t["Fin (sem.)"], "resp": t["Responsable"], "order": len(lots)}
    else:
        lots[key]["start"] = min(lots[key]["start"], t["Début (sem.)"]); lots[key]["end"] = max(lots[key]["end"], t["Fin (sem.)"])

def chunks(seq, size=120):
    for i in range(0, len(seq), size): yield seq[i:i + size]

out.append("\n-- Lots (niveau 1 du WBS) : dates = debut du projet + semaines")
lot_rows = [f"({q(code)},{q(lot)},{q(l['name'])},{int(l['start'])},{int(l['end'])},{q(l['resp'])},{l['order'] * 100})" for (code, lot), l in lots.items()]
for batch in chunks(lot_rows):
    out.append("""insert into public.tasks (project_id, wbs_code, name, status, start_date, end_date, progress, budget, responsible_role, sort_order)
select p.id, v.wbs, v.name, 'todo', p.start_date + (v.s - 1) * 7, p.start_date + (v.e - 1) * 7 + 6, 0, 0, v.resp, v.ord
from (values\n""" + ",\n".join(batch) + """) as v(code, wbs, name, s, e, resp, ord)
join public.projects p on p.code = v.code
on conflict (project_id, wbs_code) where wbs_code is not null do update set name = excluded.name, start_date = excluded.start_date, end_date = excluded.end_date, responsible_role = excluded.responsible_role, sort_order = excluded.sort_order;""")

out.append("\n-- Taches (niveau 2 du WBS)")
task_rows = []
for i, t in enumerate(tasks):
    code, lot = t["Code projet"], t["Code lot"]
    notes = " · ".join(x for x in [f"Source : {t['Source']}" if t.get("Source") else "", f"Durée : {t['Durée (sem.)']} sem." if t.get("Durée (sem.)") is not None else ""] if x)
    task_rows.append(f"({q(code)},{q(lot)},{q(t['Code tâche'])},{q(t['Nom de la tâche'])},{int(t['Début (sem.)'])},{int(t['Fin (sem.)'])},{n(t.get('HTVA'))},{n(t.get('Douanes'))},{n(t.get('TVA'))},{q(t.get('Responsable') or '')},{q(t.get('Type de lien') or '')},{q(t.get('Méthode') or '')},{q(t.get('Confiance') or '')},{lots[(code, lot)]['order'] * 100 + (i % 100) + 1},{q(notes)})")
for batch in chunks(task_rows):
    out.append("""insert into public.tasks (project_id, parent_id, wbs_code, name, status, start_date, end_date, progress, budget, customs, vat, responsible_role, link_type, estimate_method, confidence, sort_order, notes)
select p.id, l.id, v.wbs, v.name, 'todo', p.start_date + (v.s - 1) * 7, p.start_date + (v.e - 1) * 7 + 6, 0, v.htva, v.customs, v.vat, v.resp, v.link, v.method, v.conf, v.ord, v.notes
from (values\n""" + ",\n".join(batch) + """) as v(code, lot, wbs, name, s, e, htva, customs, vat, resp, link, method, conf, ord, notes)
join public.projects p on p.code = v.code
join public.tasks l on l.project_id = p.id and l.wbs_code = v.lot
on conflict (project_id, wbs_code) where wbs_code is not null do update set parent_id = excluded.parent_id, name = excluded.name, start_date = excluded.start_date, end_date = excluded.end_date,
  budget = excluded.budget, customs = excluded.customs, vat = excluded.vat, responsible_role = excluded.responsible_role, link_type = excluded.link_type,
  estimate_method = excluded.estimate_method, confidence = excluded.confidence, sort_order = excluded.sort_order, notes = excluded.notes;""")

out.append("\n-- Dependances entre taches (FD = fin-debut, DD = debut-debut)")
dep_rows = [f"({q(t['Code projet'])},{q(t['Code tâche'])},{q(t['Tâche liée'])})" for t in tasks if t.get("Tâche liée")]
for batch in chunks(dep_rows, 300):
    out.append("""update public.tasks t set depends_on = d.id
from (values\n""" + ",\n".join(batch) + """) as v(code, wbs, dep)
join public.projects p on p.code = v.code
join public.tasks d on d.project_id = p.id and d.wbs_code = v.dep
where t.project_id = p.id and t.wbs_code = v.wbs;""")

out.append("\ncommit;")
out.append("notify pgrst, 'reload schema';")
open(OUT, "w").write("\n".join(out) + "\n")
print(f"{OUT} : {len(projects)} projets, {len(lots)} lots, {len(tasks)} taches, {len(tranches)} tranches, {sum(1 for t in tasks if t.get('Tâche liée'))} dependances")
